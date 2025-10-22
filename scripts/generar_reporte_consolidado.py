import pandas as pd
import sqlite3
import sys
import os
from openpyxl import load_workbook

# Configuración desde variables de entorno o valores por defecto
MES_TRABAJO = os.environ.get('MES_TRABAJO', 'SEPTIEMBRE')
MES_CORTO = os.environ.get('MES_CORTO', 'Sep')
ANIO = int(os.environ.get('ANIO', '2025'))

# Rutas desde variables de entorno o valores por defecto
BD_FORMACION = os.environ.get('BD_FORMACION', r'C:\ws\sena\data\PE-04\sena_formacion_septiembre.db')
CUPOS_DISPONIBLES = os.environ.get('CUPOS_DISPONIBLES', r'C:\ws\sena\data\metas\cupos_disponibles_por_regional_2025.xlsx')
REPORTE_APRENDICES = os.environ.get('REPORTE_APRENDICES', rf'C:\ws\sena\data\aprendices\SENA Mensual Nacional {MES_CORTO} {ANIO}.xlsx')
ARCHIVO_SALIDA = os.environ.get('ARCHIVO_SALIDA', rf'C:\ws\sena\data\REPORTE_ECONOMIA_NARANJA\Reporte Consolidado Economía Naranja {MES_CORTO} {ANIO}.xlsx')

# Directorios
dir_base = r'C:\ws\sena\data'
dir_economia_naranja = os.path.join(dir_base, 'REPORTE_ECONOMIA_NARANJA')

print("=== GENERACION REPORTE CONSOLIDADO ECONOMIA NARANJA ===\n")

# Función para validar longitud de nombre de hoja (máximo 31 caracteres en Excel)
def validar_nombre_hoja(nombre, max_length=31):
    """Valida y ajusta el nombre de la hoja para que cumpla con el límite de Excel"""
    if len(nombre) <= max_length:
        return nombre

    # Si es muy largo, intentar recortar el mes
    if "Economía Naranja" in nombre:
        # Probar con diferentes longitudes del mes
        for long_mes in range(len(MES_TRABAJO), 2, -1):
            mes_recortado = MES_TRABAJO[:long_mes]
            nombre_nuevo = f"Economía Naranja {mes_recortado} {ANIO}"
            if len(nombre_nuevo) <= max_length:
                print(f"   [INFO] Nombre de hoja recortado: '{nombre}' -> '{nombre_nuevo}'")
                return nombre_nuevo
        # Si aún es muy largo, usar abreviatura
        return f"Eco Naranja {MES_CORTO} {ANIO}"

    elif "Oferta Disponible" in nombre:
        for long_mes in range(len(MES_TRABAJO), 2, -1):
            mes_recortado = MES_TRABAJO[:long_mes]
            nombre_nuevo = f"Oferta Disponible {mes_recortado} {ANIO}"
            if len(nombre_nuevo) <= max_length:
                print(f"   [INFO] Nombre de hoja recortado: '{nombre}' -> '{nombre_nuevo}'")
                return nombre_nuevo
        return f"Oferta Disp {MES_CORTO} {ANIO}"

    elif "SENA Mensual Nacional" in nombre:
        return f"SENA Mensual Nacional {MES_CORTO} {ANIO}"

    # Si no coincide con ningún patrón, truncar
    return nombre[:max_length]

# Validar nombres de hojas
nombre_hoja1 = validar_nombre_hoja(f"Economía Naranja {MES_TRABAJO} {ANIO}")
nombre_hoja2 = validar_nombre_hoja(f"Oferta Disponible {MES_TRABAJO} {ANIO}")
nombre_hoja3 = f"SENA Mensual Nacional {MES_CORTO} {ANIO}"

print(f"Mes de trabajo: {MES_TRABAJO}")
print(f"Año: {ANIO}")
print(f"\nNombres de hojas:")
print(f"  1. {nombre_hoja1} (longitud: {len(nombre_hoja1)})")
print(f"  2. {nombre_hoja2} (longitud: {len(nombre_hoja2)})")
print(f"  3. {nombre_hoja3} (longitud: {len(nombre_hoja3)})")

# ============================================
# HOJA 1: ECONOMÍA NARANJA
# ============================================
print(f"\n1. Generando hoja '{nombre_hoja1}'...")

# Base de datos de economía naranja (usar la BD_FORMACION de las variables de entorno)
db_economia_naranja = BD_FORMACION

if os.path.exists(db_economia_naranja):
    print(f"   Conectando a base de datos: {db_economia_naranja}")
    try:
        conn_eco = sqlite3.connect(db_economia_naranja)

        # Verificar si existe la tabla precalculada
        cursor_eco = conn_eco.cursor()
        nombre_tabla = f"ECONOMIA_NARANJA_{MES_TRABAJO}_{ANIO}"
        cursor_eco.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{nombre_tabla}'")
        tabla_existe = cursor_eco.fetchone()

        if tabla_existe:
            print(f"   Usando tabla precalculada: {nombre_tabla}")
            consulta_simple = f"SELECT * FROM {nombre_tabla}"
            df_economia_naranja = pd.read_sql_query(consulta_simple, conn_eco)
            print(f"   [OK] {len(df_economia_naranja)} registros obtenidos")
        else:
            print(f"   [ERROR] Tabla {nombre_tabla} no encontrada en la base de datos")
            print(f"   Por favor ejecute primero el paso 4 para crear la tabla precalculada")
            conn_eco.close()
            sys.exit(1)

        conn_eco.close()
    except Exception as e:
        print(f"   [ERROR] Error al ejecutar consulta: {e}")
        df_economia_naranja = pd.DataFrame(columns=[
            'CODIGO_NIVEL_FORMACION', 'nombre_departamento', 'nombre_programa',
            'COMPLEMENTARIA', 'TITULADA', 'TOTAL'
        ])
else:
    print(f"   [ERROR] No se encontró la base de datos: {db_economia_naranja}")
    df_economia_naranja = pd.DataFrame(columns=[
        'CODIGO_NIVEL_FORMACION', 'nombre_departamento', 'nombre_programa',
        'COMPLEMENTARIA', 'TITULADA', 'TOTAL'
    ])

# ============================================
# HOJA 2: OFERTA DISPONIBLE (desde cupos disponibles)
# ============================================
print(f"\n2. Generando hoja '{nombre_hoja2}'...")

# Usar archivo generado por cruce_metas_avance_final.py (desde variable de entorno)
archivo_cupos_disponibles = CUPOS_DISPONIBLES

if not os.path.exists(archivo_cupos_disponibles):
    print(f"   [ERROR] No se encontró el archivo: {archivo_cupos_disponibles}")
    print(f"   [INFO] Ejecute primero: python metas/cruce_metas_avance_final.py")
    sys.exit(1)

print(f"   Leyendo archivo: {os.path.basename(archivo_cupos_disponibles)}")
try:
    df_oferta = pd.read_excel(archivo_cupos_disponibles, sheet_name='Cupos Disponibles')
    print(f"   [OK] {len(df_oferta)} registros obtenidos")
except Exception as e:
    print(f"   [ERROR] Error al leer archivo: {e}")
    sys.exit(1)

# ============================================
# HOJA 3: SENA MENSUAL NACIONAL (desde aprendices)
# ============================================
print(f"\n3. Preparando hoja '{nombre_hoja3}'...")

# Usar reporte de aprendices desde variable de entorno
archivo_aprendices = REPORTE_APRENDICES

if not os.path.exists(archivo_aprendices):
    print(f"   [ERROR] No se encontró el archivo: {archivo_aprendices}")
    sys.exit(1)

print(f"   [OK] Archivo encontrado: {os.path.basename(archivo_aprendices)}")

# ============================================
# GENERAR ARCHIVO CONSOLIDADO
# ============================================
print(f"\n4. Generando archivo consolidado...")

# Usar ruta de salida desde variable de entorno
ruta_salida = ARCHIVO_SALIDA

# Crear el Excel con las primeras dos hojas usando pandas
with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
    df_economia_naranja.to_excel(writer, sheet_name=nombre_hoja1, index=False)
    df_oferta.to_excel(writer, sheet_name=nombre_hoja2, index=False)

# Copiar la hoja 3 con formato desde el archivo de aprendices
print(f"\n5. Copiando hoja de aprendices con formato...")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from copy import copy

# Cargar ambos workbooks
wb_source = load_workbook(archivo_aprendices)
wb_dest = load_workbook(ruta_salida)

# Obtener la hoja fuente (primera hoja del archivo de aprendices)
ws_source = wb_source.worksheets[0]

# Crear nueva hoja en destino
ws_dest = wb_dest.create_sheet(nombre_hoja3)

# Copiar todas las celdas con su formato
for row in ws_source.iter_rows():
    for cell in row:
        new_cell = ws_dest[cell.coordinate]

        # Copiar valor
        new_cell.value = cell.value

        # Copiar formato
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

# Copiar anchos de columnas
for col_letter, col_dim in ws_source.column_dimensions.items():
    ws_dest.column_dimensions[col_letter].width = col_dim.width

# Copiar altos de filas
for row_num, row_dim in ws_source.row_dimensions.items():
    ws_dest.row_dimensions[row_num].height = row_dim.height

print(f"   [OK] Hoja copiada con formato preservado")

# ============================================
# APLICAR FORMATO A ENCABEZADOS DE LAS TRES HOJAS
# ============================================
print(f"\n6. Aplicando formato a encabezados...")

# Estilos para encabezados
header_font_white = Font(bold=True, color="FFFFFF")
fill_azul_oscuro = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Hoja 1
fill_verde = PatternFill(start_color="548235", end_color="548235", fill_type="solid")  # Hojas 2 y 3

# Aplicar formato a Hoja 1 (Economía Naranja)
ws_hoja1 = wb_dest[nombre_hoja1]
for col in range(1, ws_hoja1.max_column + 1):
    cell = ws_hoja1.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = fill_azul_oscuro
print(f"   [OK] Formato aplicado a '{nombre_hoja1}' (azul #1F4E78)")

# Aplicar formato a Hoja 2 (Oferta Disponible)
ws_hoja2 = wb_dest[nombre_hoja2]
for col in range(1, ws_hoja2.max_column + 1):
    cell = ws_hoja2.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = fill_verde
print(f"   [OK] Formato aplicado a '{nombre_hoja2}' (verde #548235)")

# Aplicar formato a Hoja 3 (SENA Mensual Nacional)
for col in range(1, ws_dest.max_column + 1):
    cell = ws_dest.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = fill_verde
print(f"   [OK] Formato aplicado a '{nombre_hoja3}' (verde #548235)")

# Guardar archivo final
wb_dest.save(ruta_salida)
wb_source.close()
wb_dest.close()

print(f"   [OK] Todos los formatos aplicados correctamente")

print(f"\n[OK] Reporte consolidado generado exitosamente:")
print(f"   {ruta_salida}")

# ============================================
# ESTADÍSTICAS
# ============================================
print(f"\n6. Estadísticas del reporte:")
print(f"   Hoja 1 '{nombre_hoja1}': {len(df_economia_naranja)} registros")
print(f"   Hoja 2 '{nombre_hoja2}': {len(df_oferta)} registros")
print(f"   Hoja 3 '{nombre_hoja3}': Copiada desde {os.path.basename(archivo_aprendices)} (con formato)")

print("\nProceso completado exitosamente!")
