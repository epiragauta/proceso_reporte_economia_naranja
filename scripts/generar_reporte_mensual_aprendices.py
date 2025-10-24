import pandas as pd
import re
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import sys

# Configuración
if len(sys.argv) < 2:
    print("Uso: python generar_reporte_mensual_aprendices.py <ruta_archivo_entrada>")
    exit(1)
archivo_entrada = sys.argv[1]

print("=== GENERACION REPORTE MENSUAL NACIONAL SENA ===\n")

# 1. EXTRAER MES Y AÑO DEL NOMBRE DEL ARCHIVO
print("1. Extrayendo información del archivo...")
nombre_archivo = Path(archivo_entrada).name.strip()
print(f"   Archivo: {nombre_archivo}")

# Buscar mes y año en el nombre del archivo
meses = {
    "ENERO": "Ene",
    "FEBRERO": "Feb",
    "MARZO": "Mar",
    "ABRIL": "Abr",
    "MAYO": "May",
    "JUNIO": "Jun",
    "JULIO": "Jul",
    "AGOSTO": "Ago",
    "SEPTIEMBRE": "Sep",
    "OCTUBRE": "Oct",
    "NOVIEMBRE": "Nov",
    "DICIEMBRE": "Dic",
}

meses_numeros = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

mes_nombre = None
mes_corto = None
mes_numero = None
anio = None

for mes_completo, mes_abr in meses.items():
    if mes_completo in nombre_archivo.upper():
        mes_nombre = mes_completo
        mes_corto = mes_abr
        mes_numero = meses_numeros[mes_completo]
        break

# Buscar año (4 dígitos)
match_anio = re.search(r"20\d{2}", nombre_archivo)
if match_anio:
    anio = int(match_anio.group())

if not mes_nombre or not anio:
    print("   [ERROR] No se pudo extraer mes y año del nombre del archivo")
    exit(1)

# Calcular último día del mes
ultimo_dia = calendar.monthrange(anio, mes_numero)[1]
fecha_corte = f"{ultimo_dia} de {mes_nombre.capitalize()} {anio}"

print(f"   Mes: {mes_nombre} ({mes_corto})")
print(f"   Año: {anio}")
print(f"   Fecha de corte: {fecha_corte}")


# 2. FUNCIÓN PARA DETECTAR FILA DE HEADER
def detectar_header(df, palabras_clave=["DEPARTAMENTO", "DEPTO", "MUNICIPIO", "MPIO"]):
    """Detecta automáticamente la fila del header"""
    for i in range(min(20, len(df))):
        fila_str = " ".join(
            [str(x).upper() for x in df.iloc[i].tolist() if pd.notna(x)]
        )
        if any(palabra in fila_str for palabra in palabras_clave):
            return i
    return None


# 3. LEER DATOS DE CADA HOJA
print("\n2. Leyendo datos de cada hoja...")

# 3.1 INTEGRACION DEPTO MPIO - Columna F (Doble Titulación)
print("   - Leyendo INTEGRACION DEPTO MPIO...")
df_temp = pd.read_excel(
    archivo_entrada, sheet_name="INTEGRACION DEPTO MPIO ", engine="pyxlsb", header=None
)
header_row = detectar_header(df_temp)
df_integracion = pd.read_excel(
    archivo_entrada,
    sheet_name="INTEGRACION DEPTO MPIO ",
    engine="pyxlsb",
    header=header_row,
)
# Columna F es índice 5, pero después del header es columna 8 (TOTAL)
df_doble_tit = df_integracion.iloc[:, [0, 1, 2, 3, 8]].copy()
df_doble_tit.columns = [
    "codigo_depto",
    "nombre_depto",
    "codigo_mpio",
    "nombre_mpio",
    "doble_titulacion",
]

# 3.2 2025 DPTO_MPIO GENERO - Columnas AK, AO, AS
print("   - Leyendo 2025 DPTO_MPIO GENERO...")
df_temp = pd.read_excel(
    archivo_entrada, sheet_name="2025 DPTO_MPIO  GENERO ", engine="pyxlsb", header=None
)
header_row = detectar_header(df_temp)
df_genero = pd.read_excel(
    archivo_entrada,
    sheet_name="2025 DPTO_MPIO  GENERO ",
    engine="pyxlsb",
    header=header_row,
)

# AK = columna 36, AO = columna 40, AS = columna 44
# Necesitamos buscar por nombre de columna que contenga "TOTAL"
col_titulada = None
col_complementaria = None
col_integral = None

for i, col in enumerate(df_genero.columns):
    col_str = str(col).upper()
    if "FORMACION TITULADA" in col_str and "TOTAL" in col_str:
        col_titulada = i
    elif "FORMACION COMPLEMENTARIA" in col_str and "TOTAL" in col_str:
        col_complementaria = i
    elif "GRAN TOTAL" in col_str and col_integral is None:
        col_integral = i

if col_titulada is None or col_complementaria is None or col_integral is None:
    print("   [ADVERTENCIA] Columnas no encontradas por nombre, usando índices fijos")
    col_titulada = 36
    col_complementaria = 40
    col_integral = 44

df_formacion = df_genero.iloc[
    :, [0, 1, 2, 3, col_titulada, col_complementaria, col_integral]
].copy()
df_formacion.columns = [
    "codigo_depto",
    "nombre_depto",
    "codigo_mpio",
    "nombre_mpio",
    "formacion_titulada",
    "formacion_complementaria",
    "formacion_integral",
]

# 3.3 DEPTO MPIO VIRTUAL - Columna I
print("   - Leyendo DEPTO MPIO VIRTUAL...")
df_temp = pd.read_excel(
    archivo_entrada, sheet_name="DEPTO MPIO VIRTUAL", engine="pyxlsb", header=None
)
header_row = detectar_header(df_temp)
df_virtual = pd.read_excel(
    archivo_entrada, sheet_name="DEPTO MPIO VIRTUAL", engine="pyxlsb", header=header_row
)
df_virtualidad = df_virtual.iloc[:, [0, 1, 2, 3, 8]].copy()
df_virtualidad.columns = [
    "codigo_depto",
    "nombre_depto",
    "codigo_mpio",
    "nombre_mpio",
    "virtualidad",
]

# 3.4 DEPTO MPIO BILINGUISMO - Columna I
print("   - Leyendo DEPTO MPIO BILINGUISMO...")
df_temp = pd.read_excel(
    archivo_entrada, sheet_name="DEPTO MPIO BILINGUISMO", engine="pyxlsb", header=None
)
header_row = detectar_header(df_temp)
df_bilin = pd.read_excel(
    archivo_entrada,
    sheet_name="DEPTO MPIO BILINGUISMO",
    engine="pyxlsb",
    header=header_row,
)
df_bilinguismo = df_bilin.iloc[:, [0, 1, 2, 3, 8]].copy()
df_bilinguismo.columns = [
    "codigo_depto",
    "nombre_depto",
    "codigo_mpio",
    "nombre_mpio",
    "bilinguismo",
]

# 3.5 POBL. VULN DEPTO MPIO - Columnas AB, AP, BC, BG, AZ
print("   - Leyendo POBL. VULN DEPTO MPIO...")
df_temp = pd.read_excel(
    archivo_entrada, sheet_name="POBL. VULN DEPTO MPIO ", engine="pyxlsb", header=None
)
header_row = detectar_header(df_temp)
df_vuln = pd.read_excel(
    archivo_entrada,
    sheet_name="POBL. VULN DEPTO MPIO ",
    engine="pyxlsb",
    header=header_row,
)

# AB=27, AP=41, BC=54, BG=58, AZ=51
# Buscar por nombre de columna
col_victimas = None
col_discapacidad = None
col_mujer_cabeza = None
col_tercera_edad = None
col_indigena = None

for i, col in enumerate(df_vuln.columns):
    col_str = str(col).upper()
    if "VICTIMA" in col_str and "TOTAL" in col_str:
        col_victimas = i
    elif "DISCAPACIDAD" in col_str and "TOTAL" in col_str:
        col_discapacidad = i
    elif "MUJER" in col_str and "CABEZA" in col_str:
        col_mujer_cabeza = i
    elif "TERCERA EDAD" in col_str or ("ADULTO" in col_str and "MAYOR" in col_str):
        col_tercera_edad = i
    elif "IND" in col_str and "GENA" in col_str:
        col_indigena = i

if (
    col_victimas is None
    or col_discapacidad is None
    or col_mujer_cabeza is None
    or col_tercera_edad is None
    or col_indigena is None
):
    print(
        "   [ADVERTENCIA] Columnas vulnerables no encontradas por nombre, usando índices fijos"
    )
    col_victimas = 27
    col_discapacidad = 41
    col_mujer_cabeza = 54
    col_tercera_edad = 58
    col_indigena = 51

df_poblacion = df_vuln.iloc[
    :,
    [
        0,
        1,
        2,
        3,
        col_victimas,
        col_discapacidad,
        col_mujer_cabeza,
        col_tercera_edad,
        col_indigena,
    ],
].copy()
df_poblacion.columns = [
    "codigo_depto",
    "nombre_depto",
    "codigo_mpio",
    "nombre_mpio",
    "victimas",
    "discapacidad",
    "mujer_cabeza_familia",
    "tercera_edad",
    "indigena",
]

print("   [OK] Todas las hojas leidas")

# 4. COMBINAR DATOS
print("\n3. Combinando datos...")


# Limpiar y preparar dataframes conservando nombres
def limpiar_df(df):
    """Limpia el dataframe eliminando filas no válidas pero conservando nombres"""
    df = df.copy()
    # Convertir códigos a numéricos
    df["codigo_depto"] = pd.to_numeric(df["codigo_depto"], errors="coerce")
    df["codigo_mpio"] = pd.to_numeric(df["codigo_mpio"], errors="coerce")
    # Eliminar filas sin código de departamento válido
    df = df[df["codigo_depto"].notna()].copy()
    df = df[df["codigo_mpio"].notna()].copy()
    # Convertir a enteros
    df["codigo_depto"] = df["codigo_depto"].astype(int)
    df["codigo_mpio"] = df["codigo_mpio"].astype(int)
    return df


df_doble_tit = limpiar_df(df_doble_tit)
df_formacion = limpiar_df(df_formacion)
df_virtualidad = limpiar_df(df_virtualidad)
df_bilinguismo = limpiar_df(df_bilinguismo)
df_poblacion = limpiar_df(df_poblacion)

# CREAR CATÁLOGO MAESTRO DE MUNICIPIOS
print("   Creando catálogo maestro de municipios desde todas las hojas...")


def crear_catalogo_maestro(dfs_list):
    """Crea un catálogo único de todos los municipios encontrados en todas las hojas"""
    catalogo = []

    for df in dfs_list:
        # Extraer combinaciones únicas de código y nombre
        municipios = df[
            ["codigo_depto", "nombre_depto", "codigo_mpio", "nombre_mpio"]
        ].copy()
        municipios = municipios.drop_duplicates()
        catalogo.append(municipios)

    # Concatenar todos
    catalogo_completo = pd.concat(catalogo, ignore_index=True)

    # Eliminar duplicados, priorizando registros con nombres no vacíos
    catalogo_completo["tiene_nombre_depto"] = catalogo_completo[
        "nombre_depto"
    ].notna() & (catalogo_completo["nombre_depto"] != "")
    catalogo_completo["tiene_nombre_mpio"] = catalogo_completo[
        "nombre_mpio"
    ].notna() & (catalogo_completo["nombre_mpio"] != "")

    # Ordenar para que registros con nombres aparezcan primero
    catalogo_completo = catalogo_completo.sort_values(
        ["codigo_depto", "codigo_mpio", "tiene_nombre_depto", "tiene_nombre_mpio"],
        ascending=[True, True, False, False],
    )

    # Eliminar duplicados quedándonos con el primero (que tiene nombre)
    catalogo_unico = catalogo_completo.drop_duplicates(
        subset=["codigo_depto", "codigo_mpio"], keep="first"
    )[["codigo_depto", "nombre_depto", "codigo_mpio", "nombre_mpio"]]

    return catalogo_unico


# Crear catálogo maestro con todas las hojas
catalogo_maestro = crear_catalogo_maestro(
    [df_doble_tit, df_formacion, df_virtualidad, df_bilinguismo, df_poblacion]
)

print(f"   Catálogo maestro creado: {len(catalogo_maestro)} municipios únicos")

# Combinar datos numéricos de cada hoja
print("   Combinando datos numéricos...")

# Extraer solo columnas numéricas de cada dataframe
num_doble_tit = df_doble_tit[["codigo_depto", "codigo_mpio", "doble_titulacion"]].copy()
num_formacion = df_formacion[
    [
        "codigo_depto",
        "codigo_mpio",
        "formacion_titulada",
        "formacion_complementaria",
        "formacion_integral",
    ]
].copy()
num_virtualidad = df_virtualidad[["codigo_depto", "codigo_mpio", "virtualidad"]].copy()
num_bilinguismo = df_bilinguismo[["codigo_depto", "codigo_mpio", "bilinguismo"]].copy()
num_poblacion = df_poblacion[
    [
        "codigo_depto",
        "codigo_mpio",
        "victimas",
        "discapacidad",
        "mujer_cabeza_familia",
        "tercera_edad",
        "indigena",
    ]
].copy()

# Usar catálogo maestro como base y hacer merge con datos numéricos
resultado = catalogo_maestro.copy()
resultado = resultado.merge(
    num_doble_tit, on=["codigo_depto", "codigo_mpio"], how="left"
)
resultado = resultado.merge(
    num_formacion, on=["codigo_depto", "codigo_mpio"], how="left"
)
resultado = resultado.merge(
    num_virtualidad, on=["codigo_depto", "codigo_mpio"], how="left"
)
resultado = resultado.merge(
    num_bilinguismo, on=["codigo_depto", "codigo_mpio"], how="left"
)
resultado = resultado.merge(
    num_poblacion, on=["codigo_depto", "codigo_mpio"], how="left"
)

# Agregar código de departamento numérico para ordenar
resultado["codigo_depto_num"] = resultado["codigo_depto"]

# Llenar NaN con 0
columnas_numericas = [
    "doble_titulacion",
    "formacion_titulada",
    "formacion_complementaria",
    "formacion_integral",
    "virtualidad",
    "bilinguismo",
    "victimas",
    "discapacidad",
    "mujer_cabeza_familia",
    "tercera_edad",
    "indigena",
]

for col in columnas_numericas:
    if col in resultado.columns:
        resultado[col] = resultado[col].fillna(0).astype(int)

# 5. GENERAR CÓDIGO DIVIPOLA
print("\n4. Generando códigos DIVIPOLA...")


def generar_divipola_mpio(row):
    """Genera código DIVIPOLA a partir de código de departamento y municipio"""
    try:
        codigo_depto_str = str(int(row["codigo_depto"])).zfill(2)
        codigo_mpio_str = str(int(row["codigo_mpio"])).zfill(3)
        return f"{codigo_depto_str}{codigo_mpio_str}"
    except:
        return ""


resultado["codigo_divipola"] = resultado.apply(generar_divipola_mpio, axis=1)

# 6. CREAR REGISTROS DE DEPARTAMENTOS (AGRUPADOS)
print("\n5. Creando registros agrupados por departamento...")

# Agrupar por departamento y sumar
columnas_suma = [
    "doble_titulacion",
    "formacion_titulada",
    "formacion_complementaria",
    "formacion_integral",
    "virtualidad",
    "bilinguismo",
    "victimas",
    "discapacidad",
    "mujer_cabeza_familia",
    "tercera_edad",
    "indigena",
]

df_deptos = (
    resultado.groupby(["codigo_depto_num", "nombre_depto"])[columnas_suma]
    .sum()
    .reset_index()
)

# Generar código DIVIPOLA para departamentos (código + 000)
df_deptos["codigo_divipola"] = df_deptos["codigo_depto_num"].apply(
    lambda x: f"{str(int(x)).zfill(2)}000"
)
df_deptos["nombre_mpio"] = ""  # Departamentos no tienen municipio
df_deptos["es_departamento"] = True

# Marcar municipios
resultado["es_departamento"] = False

# 7. COMBINAR DEPARTAMENTOS Y MUNICIPIOS
print("\n6. Combinando departamentos y municipios...")

# Asegurar que ambos dataframes tengan las mismas columnas
df_deptos["codigo_depto"] = df_deptos["codigo_depto_num"]
df_deptos["codigo_mpio"] = 0  # Indicador de que es departamento

# Concatenar departamentos y municipios
df_completo = pd.concat([df_deptos, resultado], ignore_index=True)

# Ordenar por código de departamento, luego por es_departamento (True primero), luego por municipio
df_completo = df_completo.sort_values(
    ["codigo_depto_num", "es_departamento", "nombre_mpio"],
    ascending=[True, False, True],
)

# 8. PREPARAR DATAFRAME FINAL
print("\n7. Preparando reporte final...")

# Seleccionar y ordenar columnas según especificación
df_final = pd.DataFrame(
    {
        "DEPARTAMENTO": df_completo["nombre_depto"],
        "MUNICIPIO": df_completo["nombre_mpio"],
        "Código DIVIPOLA - DANE": df_completo["codigo_divipola"],
        f"Aprendices Doble Titulación - Corte: {fecha_corte}": df_completo[
            "doble_titulacion"
        ],
        f"Aprendices en formación titulada - Corte: {fecha_corte}": df_completo[
            "formacion_titulada"
        ],
        f"Aprendices en formación complementaria - Corte: {fecha_corte}": df_completo[
            "formacion_complementaria"
        ],
        f"Total aprendices en formación profesional integral - Corte: {fecha_corte}": df_completo[
            "formacion_integral"
        ],
        f"Aprendices de virtualidad - Corte: {fecha_corte}": df_completo["virtualidad"],
        f"Aprendices de bilingüismo - Corte: {fecha_corte}": df_completo["bilinguismo"],
        f"Aprendices con contrato de aprendizaje - Corte: {fecha_corte}": "",  # VACÍO
        f"Aprendices TOTAL VICTIMAS - Corte: {fecha_corte}": df_completo["victimas"],
        f"Total Aprendices con Discapacidad - Corte: {fecha_corte}": df_completo[
            "discapacidad"
        ],
        f"Aprendices Mujer Cabeza de Familia - Corte: {fecha_corte}": df_completo[
            "mujer_cabeza_familia"
        ],
        f"Aprendices Tercera Edad - Corte: {fecha_corte}": df_completo["tercera_edad"],
        f"Aprendices Indígena - Corte: {fecha_corte}": df_completo["indigena"],
        "es_departamento": df_completo["es_departamento"],
    }
)

print(f"   Total registros: {len(df_final)}")
print(f"   Departamentos: {df_final['es_departamento'].sum()}")
print(f"   Municipios: {(~df_final['es_departamento']).sum()}")

# 9. EXPORTAR RESULTADOS
print("\n8. Exportando reporte...")

import time

nombre_salida = f"SENA Mensual Nacional {mes_corto} {anio}.xlsx"
nombre_hoja = (
    f"SENA Mensual Nacional {mes_corto} {anio}"  # Nombre de la hoja sin extensión
)
nombre_temp = f"SENA Mensual Nacional {mes_corto} {anio}_temp_{int(time.time())}.xlsx"
dir_base = Path(archivo_entrada).parent.parent.parent
ruta_salida = dir_base / "aprendices" / nombre_temp
ruta_final = dir_base / "aprendices" / nombre_salida
ruta_intermedia = (
    dir_base
    / "proceso_reporte_economia_naranja"
    / "MESES"
    / mes_completo
    / "datos_intermedios"
    / nombre_salida
)
# Exportar sin la columna auxiliar
df_export = df_final.drop(columns=["es_departamento"])
df_export.to_excel(ruta_salida, index=False, sheet_name=nombre_hoja)
df_export.to_excel(ruta_intermedia, index=False, sheet_name=nombre_hoja)

# 10. APLICAR FORMATO A DEPARTAMENTOS
print("\n9. Aplicando formato a registros de departamentos...")

wb = load_workbook(ruta_salida)
ws = wb[nombre_hoja]

# Estilo para departamentos
fuente_negrita = Font(bold=True)
fondo_naranja = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

# Aplicar formato a las filas de departamentos (fila 2 en adelante, fila 1 es header)
fila_excel = 2  # Comienza en fila 2 (después del header)
for idx, es_depto in enumerate(df_final["es_departamento"]):
    if es_depto:
        # Aplicar formato a toda la fila
        for col in range(1, len(df_export.columns) + 1):
            celda = ws.cell(row=fila_excel, column=col)
            celda.font = fuente_negrita
            celda.fill = fondo_naranja
    fila_excel += 1

# Guardar cambios
wb.save(ruta_salida)

# Intentar renombrar al nombre final
try:
    import os

    if os.path.exists(ruta_final):
        os.remove(ruta_final)
    os.rename(ruta_salida, ruta_final)
    print("\n[OK] Reporte generado exitosamente:")
    print(f"   {ruta_final}")
except Exception as e:
    print("\n[ADVERTENCIA] No se pudo renombrar al archivo final (puede estar abierto)")
    print(f"   Archivo generado: {ruta_salida}")
    print(
        f"   Por favor, cierra el archivo Excel y renombra manualmente a: {nombre_salida}"
    )

# 11. MOSTRAR MUESTRA
print("\n10. Muestra del reporte (primeras 20 filas):")
print(df_export.head(20).to_string())

# 12. TOTALES
print("\n" + "=" * 120)
print("TOTALES NACIONALES")
print("-" * 120)

totales = df_export.select_dtypes(include=["int64", "float64"]).sum()
for col in df_export.columns[3:]:  # Desde la 4ta columna (datos numéricos)
    if col in totales:
        print(f"{col}: {int(totales[col]):>15,}")

print("\nProceso completado exitosamente!")
